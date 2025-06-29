# File: digitalization_opportunity_extractor.py

import fitz  # PyMuPDF
import pytesseract
import io
from PIL import Image
import pandas as pd
import asyncio
import os
import nltk
import json
import spacy
import cv2
import numpy as np
import re
from tabulate import tabulate
from tqdm import tqdm
from fpdf import FPDF
from openai import OpenAI
from difflib import SequenceMatcher
import time
import tiktoken
from openpyxl.utils import get_column_letter

from utils import (
    ask_gpt4o,
    translate_text_with_gpt,
    preprocess_image_for_ocr,
    extract_text_from_pages,
    smart_sentence_split,
    deduplicate_findings,
    translate_digitalization_findings
)

# Updated for multilingual detection → post-translation
async def run_pipeline(pdf_path, output_queue=None):
    client = OpenAI(
        api_key="")
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    chunk_size = 1000
    all_small_dfs = []
    all_findings = []
    semaphore = asyncio.Semaphore(10)

    total_tokens = 0
    total_cost = 0.0

    for chunk_start in range(0, total_pages, chunk_size):
        chunk_end = min(chunk_start + chunk_size, total_pages)
        msg1 = f"\n🚀 Processing Pages {chunk_start+1} to {chunk_end}..."
        print(msg1)
        if output_queue is not None:
            await output_queue.put(("update", msg1))

        raw_pages = extract_text_from_pages(doc, chunk_start, chunk_end)
        tasks = []

        for page_number, text in raw_pages:
            sentences = smart_sentence_split(text)
            batch_size = 15
            for i in range(0, len(sentences), batch_size):
                batch = sentences[i:i+batch_size]
                tasks.append(asyncio.create_task(ask_gpt4o(batch, page_number, semaphore, client)))

        results = []
        for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Processing Batches"):
            result = await coro
            results.append(result)

        extracted_rows = []
        for findings in results:
            for item in findings:
                extracted_rows.append(item)
                all_findings.append(item)

        if extracted_rows:
            translated_findings = await translate_digitalization_findings(extracted_rows)
            deduplicated = deduplicate_findings(translated_findings)
            deduplicated.sort(key=lambda x: x['confidence'], reverse=True)

            final_rows = []
            for item in deduplicated:
                final_rows.append([
                    item['search_key'],
                    item['page'],
                    item['evidence'],
                    item['area'],
                    item['digital_solution'],
                    item['reason'],
                    item['confidence']
                ])

            small_df = pd.DataFrame(final_rows, columns=[
                "Search Key",
                "Page Number",
                "Evidence",
                "Area",
                "Suggested Digitalization",
                "Reason",
                "Confidence Score"
            ])

            msg2 = f"🎯 Successfully Processed Pages {chunk_start+1}-{chunk_end}:"
            print("\n" + msg2)
            if output_queue is not None:
                await output_queue.put(("update", msg1 + "\n" + msg2))

            all_small_dfs.append(small_df)

        print("Waiting 5 seconds before processing next chunk...")
        await asyncio.sleep(5)

    if all_findings:
        translated_final = await translate_digitalization_findings(all_findings)
        final_unique = deduplicate_findings(translated_final)
        final_unique.sort(key=lambda x: x['confidence'], reverse=True)

        final_rows = []
        for item in final_unique:
            final_rows.append([
                item['search_key'],
                item['page'],
                item['evidence'],
                item['area'],
                item['digital_solution'],
                item['reason'],
                item['confidence']
            ])

        merged_df = pd.DataFrame(final_rows, columns=[
            "Search Key",
            "Page Number",
            "Evidence",
            "Area",
            "Suggested Digitalization",
            "Reason",
            "Confidence Score"
        ])

        print("\n🎯 Final Merged Digitalization Opportunities Table:\n")
        print(tabulate(merged_df, headers='keys', tablefmt='fancy_grid', showindex=False))

        with pd.ExcelWriter("opportunities.xlsx", engine="openpyxl") as writer:
            merged_df.to_excel(writer, index=False, sheet_name="Opportunities")
            worksheet = writer.sheets["Opportunities"]
            column_widths = [80, 12, 90, 35, 45, 100, 18]
            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = width

    else:
        print("\n⚠️ No digitalization opportunities found across the entire document.")
        empty_df = pd.DataFrame(columns=[
            "Search Key",
            "Page Number",
            "Evidence",
            "Area",
            "Suggested Digitalization",
            "Reason",
            "Confidence Score"
        ])
        print(tabulate(empty_df, headers='keys', tablefmt='fancy_grid', showindex=False))
        with pd.ExcelWriter("opportunities.xlsx", engine="openpyxl") as writer:
            empty_df.to_excel(writer, index=False, sheet_name="Opportunities")
            worksheet = writer.sheets["Opportunities"]
            column_widths = [80, 12, 90, 35, 45, 100, 18]
            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = width

    return total_tokens, total_cost
